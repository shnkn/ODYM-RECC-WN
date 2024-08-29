# -*- coding: utf-8 -*-
"""
@authors: fabio carrer
"""

"""
Contains the simplified version of the ODYM-RECC model v2.5 for the resource efficiency climate change mitigation nexus,
limiting the scope to residential buildings only.
Processes, flows, stocks and parameters have been revisied. 

dependencies:
numpy >= 1.9
scipy >= 0.14

"""

def main():
    from copy import deepcopy
    import datetime
    import getpass
    import importlib
    import logging as log
    import matplotlib.pyplot as plt
    # from matplotlib.lines import Line2D
    import numpy as np   
    import os
    import openpyxl
    import pandas as pd
    import pickle
    import pylab
    from scipy.interpolate import make_interp_spline
    import shutil
    import sys
    import time
    from tqdm import tqdm
    import uuid
    log.getLogger('matplotlib.font_manager').disabled = True    # required for preventing debugging messages in some console versions
    import RECC_Paths # Import path file
    
    ##################################
    #    Section 1)  Initialize      #
    ##################################
    # add ODYM module directory to system path
    sys.path.insert(0, os.path.join(os.path.join(RECC_Paths.odym_path,'odym'),'modules'))
    
    ### 1.1.) Read main script parameters
    __version__ = str('2.5')
    ProjectSpecs_Name_ConFile = 'RECC_Config.xlsx'
    Model_Configfile = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,ProjectSpecs_Name_ConFile), data_only=True)
    ScriptConfig = {'Model Setting': Model_Configfile['Cover'].cell(4,4).value}
    Model_Configsheet = Model_Configfile[ScriptConfig['Model Setting']]
    # Read debug modus:   
    DebugCounter = 0
    while Model_Configsheet.cell(DebugCounter+1, 3).value != 'Logging_Verbosity':
        DebugCounter += 1
    ScriptConfig['Logging_Verbosity'] = Model_Configsheet.cell(DebugCounter+1,4).value # Read loggin verbosity once entry was reached.    
    # Extract user name from main file
    ProjectSpecs_User_Name = getpass.getuser()
    
    # Import packages whose location is now on the system path:    
    import ODYM_Classes as msc # import the ODYM class file
    importlib.reload(msc)
    import ODYM_Functions as msf  # import the ODYM function file
    importlib.reload(msf)
    import dynamic_stock_model as dsm # import the dynamic stock model library
    importlib.reload(dsm)
    
    Name_Script        = Model_Configsheet.cell(6,4).value
    if Name_Script != 'ODYM_RECC_Main':  # Name of this script must equal the specified name in the Excel config file
        raise AssertionError('Fatal: The name of the current script does not match to the sript name specfied in the project configuration file. Exiting the script.')
    # The model will terminate if the name of the script that is run is not identical to the script name specified in the config file.
    Name_Scenario            = Model_Configsheet.cell(7,4).value # Regional scope as torso for scenario name
    StartTime                = datetime.datetime.now()
    TimeString               = str(StartTime.year) + '_' + str(StartTime.month) + '_' + str(StartTime.day) + '__' + str(StartTime.hour) + '_' + str(StartTime.minute) + '_' + str(StartTime.second)
    ProjectSpecs_Path_Result = os.path.join(RECC_Paths.results_path, Name_Scenario + '__' + TimeString )
    
    if not os.path.exists(ProjectSpecs_Path_Result): # Create model run results directory.
        os.makedirs(ProjectSpecs_Path_Result)
    # Initialize logger
    if ScriptConfig['Logging_Verbosity'] == 'DEBUG':
        log_verbosity = eval("log.DEBUG")  
    log_filename = Name_Scenario + '__' + TimeString + '.md'
    [Mylog, console_log, file_log] = msf.function_logger(log_filename, ProjectSpecs_Path_Result,
                                                         log_verbosity, log_verbosity)
    # log header and general information
    Time_Start = time.time()
    ScriptConfig['Current_UUID'] = str(uuid.uuid4())
    Mylog.info('# Simulation from ' + time.asctime())
    Mylog.info('Unique ID of scenario run: ' + ScriptConfig['Current_UUID'])
    
    ### 1.2) Read model control parameters
    Mylog.info('### Read model control parameters')
    # Read control and selection parameters into dictionary
    ScriptConfig = msf.ParseModelControl(Model_Configsheet,ScriptConfig)
    
    Mylog.info('Script: ' + Name_Script + '.py')
    Mylog.info('Model script version: ' + __version__)
    Mylog.info('Model functions version: ' + msf.__version__())
    Mylog.info('Model classes version: ' + msc.__version__())
    Mylog.info('Current User: ' + ProjectSpecs_User_Name)
    Mylog.info('Current Scenario: ' + Name_Scenario)
    Mylog.info(ScriptConfig['Description'])
    Mylog.debug('----\n')
    
    ### 1.3) Organize model output folder and logger
    Mylog.info('### Organize model output folder and logger')
    # Copy Config file and model script into that folder
    shutil.copy(os.path.join(RECC_Paths.data_path,ProjectSpecs_Name_ConFile), os.path.join(ProjectSpecs_Path_Result, ProjectSpecs_Name_ConFile))
    
    ######################################################
    #     Section 2) Read classifications and data      #
    ######################################################
    Mylog.info('## Read classification items and define all classifications')
    
    ### 2.1) Read model run config data
    Mylog.info('### Read model run config data')
    class_filename       = str(ScriptConfig['Version of master classification']) + '.xlsx'
    Classfile            = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,class_filename), data_only=True)
    Classsheet           = Classfile['MAIN_Table']
    MasterClassification = msf.ParseClassificationFile_Main(Classsheet,Mylog)
        
    Mylog.info('Read and parse config table, including the model index table, from model config sheet.')
    IT_Aspects,IT_Description,IT_Dimension,IT_Classification,IT_Selector,IT_IndexLetter,PL_Names,PL_Description,PL_Version,PL_IndexStructure,PL_IndexMatch,PL_IndexLayer,PL_SubFolder,PL_ProxyCode,PL_ProcMethod,PL_UpdateOverwrite,PrL_Number,PrL_Name,PrL_Comment,PrL_Type,ScriptConfig = msf.ParseConfigFile(Model_Configsheet,ScriptConfig,Mylog)    
    
    Mylog.info('Define model classifications and select items for model classifications according to information provided by config file.')
    ModelClassification  = {} # Dict of model classifications
    for m in range(0,len(IT_Aspects)):
        ModelClassification[IT_Aspects[m]] = deepcopy(MasterClassification[IT_Classification[m]])
        EvalString = msf.EvalItemSelectString(IT_Selector[m],len(ModelClassification[IT_Aspects[m]].Items))
        if EvalString.find(':') > -1: # range of items is taken
            RangeStart = int(EvalString[0:EvalString.find(':')])
            RangeStop  = int(EvalString[EvalString.find(':')+1::])
            ModelClassification[IT_Aspects[m]].Items = ModelClassification[IT_Aspects[m]].Items[RangeStart:RangeStop]           
        elif EvalString.find('[') > -1: # selected items are taken
            ModelClassification[IT_Aspects[m]].Items = [ModelClassification[IT_Aspects[m]].Items[i] for i in eval(EvalString)]
        elif EvalString == 'all':
            None
        else:
            Mylog.error('Item select error for aspect ' + IT_Aspects[m] + ' were found in datafile.')
            break
        
    ### 2.2) Define model index table and parameter dictionary
    Mylog.info('### Define model index table and parameter dictionary')
    Model_Time_Start = int(min(ModelClassification['Time'].Items))
    Model_Time_End   = int(max(ModelClassification['Time'].Items))
    
    Mylog.info('Define index table dataframe.')
    IndexTable = pd.DataFrame({'Aspect'        : IT_Aspects,  # 'Time' and 'Element' must be present!
                               'Description'   : IT_Description,
                               'Dimension'     : IT_Dimension,
                               'Classification': [ModelClassification[Aspect] for Aspect in IT_Aspects],
                               'IndexLetter'   : IT_IndexLetter})  # Unique one letter (upper or lower case) indices to be used later for calculations.
    
    # Default indexing of IndexTable, other indices are produced on the fly
    IndexTable.set_index('Aspect', inplace=True)
    # Add indexSize to IndexTable:
    IndexTable['IndexSize'] = pd.Series([len(IndexTable.Classification[i].Items) for i in range(0, len(IndexTable.IndexLetter))], index=IndexTable.index)
    # list of the classifications used for each indexletter
    IndexTable_ClassificationNames = [IndexTable.Classification[i].Name for i in range(0, len(IndexTable.IndexLetter))]
    
    # 2.3) Read model data and parameters.
    Mylog.info('### Read model data and parameters.')
    
    ParFileName = os.path.join(RECC_Paths.data_path,'RECC_ParameterDict_' + ScriptConfig['RegionalScope'] + '.dat')
    try: # Load Pickle parameter dict to save processing time
        ParFileObject = open(ParFileName,'rb')  
        ParameterDict = pickle.load(ParFileObject)
        Mylog.info('Read model data and parameters from pickled file with pickle file /parameter reading sequence UUID ' + ParameterDict['Checkkey'])
        #for individual parameters load new data if specified accordingly in config file
        mo_start = 0 # set mo for re-reading a certain parameter ScriptConfig['RegionalScope']
        mo_reading_true = 0 
        for mo in range(mo_start,len(PL_Names)):
            if PL_UpdateOverwrite[mo] == 'True': # new data is supposed to be used to replace data loaded from parameter dict
                mo_reading_true += 1
                if mo_reading_true == 1:
                    Mylog.info('Updating and overwriting parameter data in pickled parameter dict for selected parameters as specified in config file:')
                if PL_SubFolder[mo] == 'default': # path is not in subfolder but in main data directory
                    ParPath = os.path.join(RECC_Paths.data_path, PL_Names[mo] + '_' + PL_Version[mo])
                else: # parameter file is in subfolder, add this to path
                    ParPath = os.path.join(RECC_Paths.data_path, PL_SubFolder[mo], PL_Names[mo] + '_' + PL_Version[mo])
                Mylog.info('Reading parameter ' + PL_Names[mo] + ' and overwriting values in pickled parameter dict')
                MetaData, Values = msf.ReadParameterXLSX(ParPath, PL_Names[mo], PL_IndexStructure[mo], PL_IndexMatch[mo],
                                                     PL_IndexLayer[mo], PL_ProcMethod[mo], MasterClassification, IndexTable,
                                                     IndexTable_ClassificationNames, ScriptConfig, Mylog, False)
                ParameterDict[PL_Names[mo]] = msc.Parameter(Name=MetaData['Dataset_Name'], ID=MetaData['Dataset_ID'],
                                                            UUID=MetaData['Dataset_UUID'], P_Res=None, MetaData=MetaData,
                                                            Indices=PL_IndexStructure[mo], Values=Values, Uncert=None,
                                                            Unit=MetaData['Dataset_Unit'])
                Mylog.info('Current parameter file UUID: ' + MetaData['Dataset_UUID'])
                Mylog.info('_')        
        Mylog.info('Reading of parameters finished.')
        Mylog.info(str(mo_reading_true) + ' parameter file(s) read additionally and overwritten in pickled parameter dict.')
        if mo_reading_true > 0: #if new parameter values were added to parameter dict from previous run
            CheckKey = str(uuid.uuid4()) # generate UUID for this parameter reading sequence.
            Mylog.info('New parameter reading sequence UUID: ' + CheckKey)
            Mylog.info('Entire parameter set stored under this UUID, will be reloaded for future calculations.')
            ParameterDict['Checkkey'] = CheckKey
            # Save to pickle file for next model run
            ParFileObject = open(ParFileName,'wb') 
            pickle.dump(ParameterDict,ParFileObject)
        else: #if no new parameter data was read
            Mylog.info('Model data and parameters were read from pickled file with pickle file /parameter reading sequence UUID ' + ParameterDict['Checkkey'])
        ParFileObject.close()      
    except:
        msf.check_dataset(RECC_Paths.data_path,PL_Names,PL_Version,PL_SubFolder,Mylog)
        ParameterDict = {}
        mo_start = 0 # set mo for re-reading a certain parameter
        for mo in range(mo_start,len(PL_Names)):
            if PL_SubFolder[mo] == 'default': # path is not in subfolder
                ParPath = os.path.join(RECC_Paths.data_path, PL_Names[mo] + '_' + PL_Version[mo])
            else: # parameter file is in subfolder, add this to path
                ParPath = os.path.join(RECC_Paths.data_path, PL_SubFolder[mo], PL_Names[mo] + '_' + PL_Version[mo])
            Mylog.info('Reading parameter ' + PL_Names[mo])
            MetaData, Values = msf.ReadParameterXLSX(ParPath, PL_Names[mo], PL_IndexStructure[mo], PL_IndexMatch[mo],
                                                 PL_IndexLayer[mo], PL_ProcMethod[mo], MasterClassification, IndexTable,
                                                 IndexTable_ClassificationNames, ScriptConfig, Mylog, False)
            ParameterDict[PL_Names[mo]] = msc.Parameter(Name=MetaData['Dataset_Name'], ID=MetaData['Dataset_ID'],
                                                        UUID=MetaData['Dataset_UUID'], P_Res=None, MetaData=MetaData,
                                                        Indices=PL_IndexStructure[mo], Values=Values, Uncert=None,
                                                        Unit=MetaData['Dataset_Unit'])
            Mylog.info('Current parameter file UUID: ' + MetaData['Dataset_UUID'])
            Mylog.info('_')
        Mylog.info('Reading of parameters finished.')
        CheckKey = str(uuid.uuid4()) # generate UUID for this parameter reading sequence.
        Mylog.info('Current parameter reading sequence UUID: ' + CheckKey)
        Mylog.info('Entire parameter set stored under this UUID, will be reloaded for future calculations.')
        ParameterDict['Checkkey'] = CheckKey
        # Save to pickle file for next model run
        ParFileObject = open(ParFileName,'wb') 
        pickle.dump(ParameterDict,ParFileObject)   
        ParFileObject.close()
        
    Mylog.debug('----\n')
    
    #############################################
    #     Section 3)  Pre-run processing        #
    #############################################
    Mylog.info('Pre-processing of data.')
    Sector_reb_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('residential buildings')
    try:
        Sector_reb_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items]
    except:
        raise AssertionError('Fatal: All selected items for aspect B must also be selected for aspect g. Exiting the script.')
    
    # 3.1) Define shortcuts for the most important index sizes amd specific indices:    
    Nt = len(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items)
    Ne = len(IndexTable.Classification[IndexTable.index.get_loc('Element')].Items)
    Nc = len(IndexTable.Classification[IndexTable.index.get_loc('Cohort')].Items)
    Nr = len(IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items)
    NS = len(IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items)
    NR = len(IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    Nm = len(IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items)
    Nn = len(IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items)
    Nx = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('x')].Items)   
    NV = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('V')].Items)
    NG = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('G')].Items)
    Ng = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items)
    NB = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items)
    No = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('o')].Items)
        
    Carbon_loc    = IndexTable.Classification[IndexTable.index.get_loc('Element')].Items.index('C')
    Electric_loc  = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('electricity')
    all_loc       = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('all')
    GWP100_loc    = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('GWP100')
    Land_loc      = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Land occupation (LOP)')
    Water_loc     = IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items.index('Water consumption potential (WCP)')
    Heating_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Heating')
    Cooling_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Cooling')
    DomstHW_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('DHW')
    Equipment_loc = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Equipment')
    Lighting_loc  = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Lighting')
    LED_loc       = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items.index('LED')
    SwitchTime = Nc-Nt+1 # Index of first model year (2016)
    IsClose_Remainder_Small = 1e-15 
    IsClose_Remainder_Large = 1e-7 
    DPI_RES        = ScriptConfig['Plot1Max'] # 100 for overview or 500 for paper plots, defined in ModelConfig_List
    
    # 3.2) Set reference population dataset.
    ParameterDict['2_P_Population_Reference'] = msc.Parameter(Name='2_P_Population_Reference', ID='2_P_Population_Reference',
                                                UUID=ParameterDict[ScriptConfig['Population_Reference']].UUID, P_Res=None, MetaData=ParameterDict[ScriptConfig['Population_Reference']].MetaData,
                                                Indices=ParameterDict[ScriptConfig['Population_Reference']].Indices, Values=ParameterDict[ScriptConfig['Population_Reference']].Values, Uncert=None,
                                                Unit=ParameterDict[ScriptConfig['Population_Reference']].MetaData['Dataset_Unit'])
    
    # 3.3) Calibrate building stock. Exctract first Population value and FApC to calculate the expected stock
    for r in range(Nr):
        pop_2015 = ParameterDict['2_P_Population_Reference'].Values[0,0,r,0]
        FApC_2015 = ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings'].Values[0,0,0,r]
        expected_stock = pop_2015 * FApC_2015
        data_stock =  ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,r].sum()
        ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,r] = ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,r] / data_stock*expected_stock
    
    # 3.4) Compile parameter for building energy conversion efficiency:
    ParameterDict['4_TC_ResidentialEnergyEfficiency'] = msc.Parameter(Name='4_TC_ResidentialEnergyEfficiency', ID='4_TC_ResidentialEnergyEfficiency',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VRrntS', Values=np.zeros((NV,NR,Nr,Nn,Nt,NS)), Uncert=None, Unit='1')
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values                                   = np.einsum('VRrn,tS->VRrntS',ParameterDict['4_TC_ResidentialEnergyEfficiency_Default'].Values[:,:,:,:,0],np.ones((Nt,NS)))
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[Heating_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Heating'].Values[Heating_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[Cooling_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Cooling'].Values[Cooling_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[DomstHW_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Heating'].Values[DomstHW_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[Equipment_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Heating'].Values[Heating_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[Lighting_loc,:,:,Electric_loc,:,:]  = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Cooling'].Values[Cooling_loc,:,:,Electric_loc,:,:] / 100
     
    # 3.5) Derive energy supply multipliers for buildings for future age-cohorts
    # From energy carrier split and conversion efficiency, the multipliers converting 1 MJ of final building energy demand into different energy carriers are determined.
    # For details around the ancillary quantity anc, see the model documentation.
    Divisor = ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values #VRrntS
    Anc = np.divide(np.einsum('VRrnt,S->VRrntS',ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values, np.ones(NS)), Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
    
    # Define energy carrier split for useful energy
    ParameterDict['3_SHA_EnergyCarrierSplit_Buildings_uf'] = msc.Parameter(Name='3_SHA_EnergyCarrierSplit_Buildings_uf', ID='3_SHA_EnergyCarrierSplit_Buildings_uf',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VRrntS', Values=np.zeros((NV,NR,Nr,Nn,Nt,NS)), Uncert=None, Unit='1')
    ParameterDict['3_SHA_EnergyCarrierSplit_Buildings_uf'].Values = np.divide(Anc, np.einsum('VRrtS,n->VRrntS',np.einsum('VRrntS->VRrtS',Anc),np.ones(Nn)), out=np.zeros_like(Divisor), where=Divisor!=0)
    
    
    # 3.6) Determine future energy intensity and material composition of residential buildings by mixing archetypes:
    # Expand building light-weighting split to all building types:
    ParameterDict['3_SHA_LightWeighting_Buildings'].Values = np.einsum('B,rtS->BrtS',np.ones(NB),ParameterDict['3_SHA_LightWeighting_Buildings'].Values[Sector_reb_loc,:,:,:]).copy()
    if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'False': # no additional lightweighting trough material substitution.
        ParameterDict['3_SHA_LightWeighting_Buildings'].Values = np.einsum('BrS,t->BrtS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values[:,:,0,:],np.ones((Nt)))
    if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'False': # no lightweighting trough UsingLessMaterialByDesign.
        ParameterDict['3_SHA_DownSizing_Buildings'].Values = np.einsum('urS,t->urtS',ParameterDict['3_SHA_DownSizing_Buildings'].Values[:,:,0,:],np.ones((Nt)))
    ParameterDict['3_MC_RECC_Buildings_RECC'] = msc.Parameter(Name='3_MC_RECC_Buildings_RECC', ID='3_MC_RECC_Buildings_RECC',
                                            UUID=None, P_Res=None, MetaData=None,
                                            Indices='cmBrS', Values=np.zeros((Nc,Nm,NB,Nr,NS)), Uncert=None,
                                            Unit='kg/m2')
    ParameterDict['3_MC_RECC_Buildings_RECC'].Values[0:115,:,:,:,:] = np.einsum('cmBr,S->cmBrS',ParameterDict['3_MC_RECC_Buildings'].Values[0:115,:,:,:],np.ones(NS))
    
    # Mix future archetypes for material composition                                                                                                                                                                            # [indexes of RES2.1+RES2.2 archetypes]
                                                                                                                                                                                                                                # [indexes of RES2.1 archetypes]
                                                                                                                                                                                                                                # [indexes of RES2.2 archetypes]
                                                                                                                                                                                                                                # [indexes of RES0 archetypes]
    ParameterDict['3_MC_RECC_Buildings_RECC'].Values[115::,:,:,:,:] = \
    np.einsum('BrcS,BmrcS->cmBrS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcS,Brm->BmrcS',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_MC_BuildingArchetypes'].Values[[24,25,26,27,28,29,30,31],:,:])) +\
    np.einsum('BrcS,BmrcS->cmBrS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcS,Brm->BmrcS',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_MC_BuildingArchetypes'].Values[[8,9,10,11,13,14,14,15],:,:])) +\
    np.einsum('BrcS,BmrcS->cmBrS',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcS,Brm->BmrcS',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_MC_BuildingArchetypes'].Values[[16,17,18,19,20,21,22,23],:,:])) +\
    np.einsum('BrcS,BmrcS->cmBrS',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcS,Brm->BmrcS',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_MC_BuildingArchetypes'].Values[[0,1,2,3,4,5,6,7],:,:]))
    # Mix future archetypes for energy carrier
                                                                                                                                                                                                                                    # Same indexing as above
    ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[115::,:,:,:,:,:] = \
    np.einsum('BrcS,BnrVcS->cBVnrS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcS,BrVn->BnrVcS',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_EI_BuildingArchetypes'].Values[[24,25,26,27,28,29,30,31],:,:,:])) +\
    np.einsum('BrcS,BnrVcS->cBVnrS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcS,BrVn->BnrVcS',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_EI_BuildingArchetypes'].Values[[8,9,10,11,13,14,14,15],:,:,:])) +\
    np.einsum('BrcS,BnrVcS->cBVnrS',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcS,BrVn->BnrVcS',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_EI_BuildingArchetypes'].Values[[16,17,18,19,20,21,22,23],:,:,:])) +\
    np.einsum('BrcS,BnrVcS->cBVnrS',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcS,BrVn->BnrVcS',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_EI_BuildingArchetypes'].Values[[0,1,2,3,4,5,6,7],:,:,:]))
    # The archetypes report useful energy for 'all' energy carriers together! Must be split into different energy carriers.
    # Will happen below as energy carrier split and final-to-useful conversion efficieny is RCP-scenario dependent.
    
    # Define time-dependent final energy parameter:
    ParameterDict['3_EI_Products_UsePhase_resbuildings_t'] = msc.Parameter(Name='3_EI_Products_UsePhase_resbuildings_t', ID='3_EI_Products_UsePhase_resbuildings_t',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='cBVnrt', Values=np.zeros((Nc,NB,NV,Nn,Nr,Nt)), Uncert=None,
                                                Unit='MJ/m2/yr')
    ParameterDict['3_MC_RECC_Buildings_t'] = msc.Parameter(Name='3_MC_RECC_Buildings_t', ID='3_MC_RECC_Buildings_t',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='mBrctS', Values=np.zeros((Nm,NB,Nr,Nc,Nt,NS)), Uncert=None,
                                                Unit='kg/m2')    
    
    # 3.7) Extrapolate fabrication yield as 2015 values
    ParameterDict['4_PY_Manufacturing'].Values[:,:,1::,:] = np.einsum('t,mgr->mgtr',np.ones(Nt-1),ParameterDict['4_PY_Manufacturing'].Values[:,:,0,:])
    
    # 3.8) Define a multi-regional RE strategy and building renovation scaleup parameter
    # not used
    
    # 3.9) Calibrate res. building energy consumption
    # ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:115,:,[Heating_loc,Cooling_loc,DomstHW_loc],:,:,:] = ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:115,:,[Heating_loc,Cooling_loc,DomstHW_loc],:,:,:]    * np.einsum('r,cBVnS->cBVnrS',ParameterDict['6_PR_Calibration'].Values[Sector_reb_loc,:],np.ones((115,NB,3,Nn,NS)))
    
    # 3.10) Define parameter for future building stock:
    # Actual future  building stock
    ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'] = msc.Parameter(Name='2_S_RECC_FinalProducts_Future_resbuildings_act', ID='2_S_RECC_FinalProducts_Future_resbuildings_act',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='StGr', Values=np.zeros((NS,Nt,NG,Nr)), Uncert=None,
                                                Unit='m2 per person')
    ParameterDict['3_IO_Buildings_UsePhase'] = msc.Parameter(Name='3_IO_Buildings_UsePhase', ID='3_IO_Buildings_UsePhase',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='tcBVrS', Values=np.zeros((Nt,Nc,NB,NV,Nr,NS)), Uncert=None,
                                                Unit='1')
    # Historic age-cohorts:
    # ParameterDict['3_IO_Buildings_UsePhase_Historic'] is a combination of climate and socioeconomic 3_IO determinants.
    # We single out the former and keep them constant and let the socioeconomic factors change according to the '3_IO_Buildings_UsePhase_Future_...' parameters.
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,Heating_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,Heating_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating,np.ones(Nt))
    
    Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW     = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,DomstHW_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,DomstHW_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW,np.ones(Nt))
    
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,Cooling_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Cooling'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,Cooling_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling,np.ones(Nt))
    
    Par_3_IO_Buildings_UsePhase_Historic_Equipment = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,Equipment_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Equipment'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Equipment[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Equipment)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,Equipment_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Equipment,np.ones(Nt))
    
    Par_3_IO_Buildings_UsePhase_Historic_Lighting = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,Lighting_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Lighting'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Lighting[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Lighting)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,Lighting_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Lighting,np.ones(Nt))
    
    # Correct for if some of the corrections lead to IO > 1 which may be the case when hist. IO data are incomplete and thus set to 1 already.
    ParameterDict['3_IO_Buildings_UsePhase'].Values[ParameterDict['3_IO_Buildings_UsePhase'].Values > 1] = 1
    # Future age-cohorts:
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,Heating_loc,:,:]   = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[  Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,DomstHW_loc,:,:]   = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[  Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,Cooling_loc,:,:]   = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Cooling'].Values[  Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,Equipment_loc,:,:] = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Equipment'].Values[Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,Lighting_loc,:,:]  = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Lighting'].Values[ Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    
    # 3.11) Total 2015 material stock, all in Mt!
    TotalMaterialStock_2015_reb = np.einsum('cgr,cmgr->mgr',ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,:],ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,:,:,:,0])/1000
    
    # 3.12) Extrapolate process extensions
    # Replicate 2015 values for 4_PE_ProcessExtensions_Industry
    ParameterDict['4_PE_ProcessExtensions_Industry'].Values = np.einsum('Ixo,t->Ixot',ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,:,0],np.ones((Nt)))
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'] = msc.Parameter(Name='4_PE_ProcessExtensions_EnergyCarriers_MJ_r', ID='4_PE_ProcessExtensions_EnergyCarriers_MJ_r',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='nxrtR', Values=np.zeros((Nn,Nx,Nr,Nt,NR)), Uncert=None,
                                                Unit='[impact unit]/MJ')
    # Replicate 2015 values. In current dataset is only initial value. Electricity is added separately in the next step
    # Formula calculates impact / kg * kg /MJ = impact / MJ
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values = np.einsum('nx,n,rtR->nxrtR',ParameterDict['4_PE_ProcessExtensions_EnergyCarriers'].Values[:,:,0,0],ParameterDict['3_EI_SpecificEnergy_EnergyCarriers'].Values,np.ones((Nr,Nt,NR)))
    # Add electricity calculated from electriciy mix
    ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[Electric_loc,:,:,:,:] = np.einsum('rRIt,Ixt->xrtR', ParameterDict['4_SHA_ElectricityMix'].Values, ParameterDict['4_PE_ProcessExtensions_Industry'].Values[:,:,0,:]/3.6)
    
    Mylog.debug('----\n')
    
    
    ##########################################################
    #    Section 4) Initialize dynamic MFA model for RECC    #
    ##########################################################
    Mylog.info('Initialize dynamic MFA model for RECC')
    Mylog.info('Define RECC system and processes.')
    
    OutputDict = {}  # Dictionary with output variables for entire model run
    
    #Define arrays for results export, systema variables
    Population = np.zeros((Nt,NS,NR))
    
    MaterialDemand        = np.zeros((Nt,Nm,NS,NR))
    MaterialDemand_pc     = np.zeros((Nt,Nm,NS,NR))
    DemolitionWaste       = np.zeros((Nt,Nm,NS,NR))
    MaterialStock         = np.zeros((Nt,Nm,NS,NR))
    MaterialStock_pc      = np.zeros((Nt,Nm,NS,NR))
    MaterialProduction    = np.zeros((Nt,Nm,NS,NR))
    MaterialProduction_pc = np.zeros((Nt,Nm,NS,NR))
    ManufacturingScrap    = np.zeros((Nt,Nm,NS,NR))
    
    ProductInflow       = np.zeros((Nt,Ng,NS,NR))
    ProductOutflow      = np.zeros((Nt,Ng,NS,NR))
    ProductStock        = np.zeros((Nt,NS,NR))
    ProductStock_byType = np.zeros((Nt,Ng,NS,NR))
    ProductStock_pc     = np.zeros((Nt,NS,NR))
    
    EnergyCons_UsePhase_byEnergyCarrier      = np.zeros((Nt,Nn,NS,NR))
    EnergyCons_UsePhase_byEnergyCarrier_EL   = np.zeros((Nt,NS,NR))
    EnergyCons_UsePhase_byService            = np.zeros((Nt,NV,NS,NR))
    EnergyCons_Manufacturing_byEnergyCarrier = np.zeros((Nt,Nn,NS,NR))
    EnergyCons_Total_byEnergyCarrier         = np.zeros((Nt,Nn,NS,NR))
    
    Impacts_UsePhase_byEnergyCarrier          = np.zeros((Nx,Nt,Nn,NS,NR))
    Impacts_UsePhase_byEnergyCarrier_direct   = np.zeros((Nx,Nt,Nn,NS,NR))
    Impacts_UsePhase_byEnergyCarrier_indirect = np.zeros((Nx,Nt,Nn,NS,NR))
    Impacts_UsePhase_indir_EL                 = np.zeros((Nx,Nt,NS,NR))
    Impacts_UsePhase_indir_otherThanEL        = np.zeros((Nx,Nt,NS,NR))
    Impacts_UsePhase_byService                = np.zeros((Nx,Nt,NV,NS,NR))
    Impacts_UsePhase                          = np.zeros((Nx,Nt,NS,NR))
    Impacts_UsePhase_direct                   = np.zeros((Nx,Nt,NS,NR))
    Impacts_UsePhase_indirect                 = np.zeros((Nx,Nt,NS,NR))
    Impacts_Manufacturing                     = np.zeros((Nx,Nt,NS,NR))
    Impacts_Production                        = np.zeros((Nx,Nt,NS,NR))
    Impacts_Production_byMaterials            = np.zeros((Nx,Nt,Nm,NS,NR))
    Impacts_Total                             = np.zeros((Nx,Nt,NS,NR))
    
    # CO2_Uptake          = np.zeros((Nt,NS,NR))
    C_Uptake            = np.zeros((Nt,NS,NR))
    GHG_total_wo_uptake = np.zeros((Nt,NS,NR))
    # GHG_total_w_uptake  = np.zeros((Nt,NS,NR))
    
    NegInflowFlags = np.zeros((NS,NR))
    ExitFlags = {} # Exit flags for individual model runs
    
    F_1_2_aggregated = np.zeros((Nt,No,Ng,Nm,Ne))
    
    # Select and loop over scenarios
    for mS in range(0,NS):
        for mR in range(0,NR):
            SName = IndexTable.loc['Scenario'].Classification.Items[mS]
            RName = IndexTable.loc['Scenario_RCP'].Classification.Items[mR]
            Mylog.info('Computing RECC model for SSP scenario ' + SName + ' and RE scenario ' + RName + '.')
            
            # Initialize MFA system
            RECC_System = msc.MFAsystem(Name='RECC_SingleScenario',
                                        Geogr_Scope='19 regions + 1 single country', #IndexTableR.Classification[IndexTableR.set_index('IndexLetter').index.get_loc('r')].Items,
                                        Unit='Mt',
                                        ProcessList=[],
                                        FlowDict={},
                                        StockDict={},
                                        ParameterDict=ParameterDict,
                                        Time_Start=Model_Time_Start,
                                        Time_End=Model_Time_End,
                                        IndexTable=IndexTable,
                                        Elements=IndexTable.loc['Element'].Classification.Items,
                                        Graphical=None)
                                  
            # Check Validity of index tables: returns true if dimensions are OK and time index is present and element list is not empty
            RECC_System.IndexTableCheck()
            # Add processes to system
            for m in range(0, len(PrL_Number)):
                RECC_System.ProcessList.append(msc.Process(Name = PrL_Name[m], ID = PrL_Number[m]))
            # Define system variables: Flows.    
            RECC_System.FlowDict['F_0_1'] = msc.Flow(Name='primary material consumption' , P_Start = 0, P_End = 1, 
                                                      Indices = 't,m,e', Values=None, Uncert=None, 
                                                      Color = None, ID = None, UUID = None)
            RECC_System.FlowDict['F_1_0'] = msc.Flow(Name='material scrap' , P_Start = 1, P_End = 0, 
                                                      Indices = 't,m,e', Values=None, Uncert=None, 
                                                      Color = None, ID = None, UUID = None)
            RECC_System.FlowDict['F_1_2'] = msc.Flow(Name='final consumption', P_Start=1, P_End=2,
                                                     Indices='t,r,g,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            RECC_System.FlowDict['F_2_0'] = msc.Flow(Name='EoL products' , P_Start = 2, P_End = 0, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            # Define system variables: Stocks.
            RECC_System.StockDict['dS_0'] = msc.Stock(Name='System environment stock change', P_Res=0, Type=1,
                                                      Indices = 't,e', Values=None, Uncert=None,
                                                      ID=None, UUID=None)
            RECC_System.StockDict['S_2'] = msc.Stock(Name='In-use stock', P_Res=2, Type=0,
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            RECC_System.StockDict['dS_2'] = msc.Stock(Name='In-use stock change', P_Res=2, Type=1,
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            RECC_System.Initialize_StockValues() # Assign empty arrays to stocks according to dimensions.
            RECC_System.Initialize_FlowValues()  # Assign empty arrays to flows according to dimensions.
            
            
            ##########################################################
            #    Section 5) Solve dynamic MFA model for RECC         #
            ##########################################################
            Mylog.info('Solve dynamic MFA model for the RECC project for all sectors chosen.')
            Stock_Detail_UsePhase_B     = np.zeros((Nt,Nc,NB,Nr))    # Unit: million m².
            Outflow_Detail_UsePhase_B   = np.zeros((Nt,Nc,NB,Nr))    # Unit: million m².
            Inflow_Detail_UsePhase_B    = np.zeros((Nt,NB,Nr))       # Unit: million m².
            F_1_2_new                   = np.zeros((Nt,Nr,Ng,Nm,Ne)) # inflow of material in new products, Mt/yr
        
            Mylog.info('Calculate inflows and outflows for use phase, residential buildings.')
            # Determine total stock and apply stock-driven model
            SF_Array                    = np.zeros((Nc,Nc,NB,Nr)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
            #Get historic stock at end of 2015 by age-cohort, and covert unit to Buildings: million m2.
            TotalStock_UsePhase_Hist_cBr = RECC_System.ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,:]
            # Determine total future stock, product level. Units: Buildings: million m2.
            TotalStockCurves_UsePhase_B_pC_test = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings'].Values[mS,:,Sector_reb_loc,:]
                
            # Include (or not) the RE strategies for the use phase:
            # a) Include_REStrategy_MoreIntenseUse:
            if ScriptConfig['Include_REStrategy_MoreIntenseUse'] == 'True': 
                # Calculate counter-factual scenario: X% decrease of stock levels by 2050 compared to scenario reference. X coded in parameter ..._MIUPotential
                if SName != 'LED':
                    RemainingFraction = 1-RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_MIUPotential'].Values[Sector_reb_loc,0,mS] / 100
                    clamped_spline = make_interp_spline([0,2,Nt-5,Nt], [1,1,RemainingFraction,RemainingFraction], bc_type=([(2, 0)], [(1, 0)]))
                    MIURamp_Spline = clamped_spline(np.arange(0,Nt,1))
                    MIURamp_Spline[MIURamp_Spline>1]=1
                    MIURamp_Spline[MIURamp_Spline<RemainingFraction]=RemainingFraction
                    TotalStockCurves_UsePhase_B_pC_test    = TotalStockCurves_UsePhase_B_pC_test * np.einsum('t,r->tr',MIURamp_Spline,np.ones((Nr)))
            # Make sure that for no scenario, stock values are below LED values, which is assumed to be the lowest possible stock level.
            TotalStockCurves_UsePhase_B_pC_LED_ref = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings'].Values[LED_loc,:,Sector_reb_loc,:]
            TotalStockCurves_UsePhase_B_pC         = np.maximum(TotalStockCurves_UsePhase_B_pC_test,TotalStockCurves_UsePhase_B_pC_LED_ref)
            TotalStockCurves_UsePhase_B            = np.einsum('tr,tr->tr',TotalStockCurves_UsePhase_B_pC,RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS]) 
            RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'].Values[mS,:,Sector_reb_loc,:] = TotalStockCurves_UsePhase_B_pC.copy()
        
            # b) Include_REStrategy_LifeTimeExtension: Product lifetime extension.
            Par_RECC_ProductLifetime_B = RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values.copy()
            # Second, change lifetime of future age-cohorts according to lifetime extension parameter
            if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
                # gradual increase of lifetime by age-cohort, including historic age-cohorts, starting from 0:
                for B in range(0, NB):
                    for r in range(0, Nr):
                        LTE_Pot = RECC_System.ParameterDict['6_PR_LifeTimeExtension_resbuildings'].Values[B,r,mS]
                        LTE_Rampupcurve = np.zeros(Nc)
                        try:
                            LTE_Rampupcurve[0:SwitchTime] = np.arange(0,LTE_Pot,LTE_Pot/SwitchTime)
                        except:
                            None # LTE_Pot = 0, no LTE
                        LTE_Rampupcurve[SwitchTime::] = LTE_Pot
                        Par_RECC_ProductLifetime_B[B,r,:] = np.einsum('c,c->c',1 + LTE_Rampupcurve,Par_RECC_ProductLifetime_B[B,r,:])
    
            # Dynamic stock model, with lifetime depending on age-cohort.
            # Build pdf array from lifetime distribution: Probability of survival.
            for B in tqdm(range(0, NB), unit=' res. building types'):
                for r in range(0, Nr):
                    LifeTimes = Par_RECC_ProductLifetime_B[B, r, :]
                    lt = {'Type'  : 'Normal',
                          'Mean'  : LifeTimes,
                          'StdDev': 0.3 * LifeTimes}
                    SF_Array[:, :, B, r] = dsm.DynamicStockModel(t=np.arange(0, Nc, 1), lt=lt).compute_sf().copy()
                    np.fill_diagonal(SF_Array[:, :, B, r],1) # no outflows from current year, 
                    # this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known. Those parts of the stock remain in use instead.
    
            # Compute evolution of 2015 in-use stocks: initial stock evolution separately from future stock demand and stock-driven model
            for r in range(0,Nr):   
                FutureStock                 = np.zeros((Nc))
                FutureStock[SwitchTime::]   = TotalStockCurves_UsePhase_B[1::, r].copy()# Future total stock
                InitialStock                = TotalStock_UsePhase_Hist_cBr[:,:,r].copy()
                SFArrayCombined             = SF_Array[:,:,:,r]
                TypeSplit                   = np.zeros((Nc,NB))
                TypeSplit[SwitchTime::,:]   = RECC_System.ParameterDict['3_SHA_TypeSplit_Buildings'].Values[:,r,1::,mS].transpose() # indices: Bc
                
                RECC_dsm                    = dsm.DynamicStockModel(t=np.arange(0,Nc,1), s=FutureStock.copy(), lt = lt)  # The lt parameter is not used, the sf array is handed over directly in the next step.   
                Var_S, Var_O, Var_I, IFlags = RECC_dsm.compute_stock_driven_model_initialstock_typesplit_negativeinflowcorrect(SwitchTime,InitialStock,SFArrayCombined,TypeSplit,NegativeInflowCorrect = True)
                
                Stock_Detail_UsePhase_B[0,:,:,r]     += InitialStock.copy() # cgr, needed for correct calculation of mass balance later.
                Stock_Detail_UsePhase_B[1::,:,:,r]   += Var_S[SwitchTime::,:,:].copy() # tcBr
                Outflow_Detail_UsePhase_B[1::,:,:,r] += Var_O[SwitchTime::,:,:].copy() # tcBr
                Inflow_Detail_UsePhase_B[1::,:,r]    += Var_I[SwitchTime::,:].copy() # tBr
                # Check for negative inflows:
                if IFlags.sum() != 0:
                    NegInflowFlags[mS,mR] = 1 # flag this scenario
    
            # Here so far: Units: Buildings: million m². for stocks, X/yr for flows.
            Population[:,mS,mR]                         = np.einsum('tr->t', RECC_System.ParameterDict['2_P_Population_Reference'].Values[0,:,:,mS])
            ProductStock[:,mS,mR]                       = TotalStockCurves_UsePhase_B.sum(axis=1).copy()
            ProductStock_pc[:,mS,mR]                    = ProductStock[:,mS,mR] / Population[:,mS,mR]
            ProductStock_byType[:,Sector_reb_rge,mS,mR] = np.einsum('tcBr->tB',Stock_Detail_UsePhase_B).copy()
            ProductInflow[:,Sector_reb_rge,mS,mR]       = np.einsum('tBr->tB', Inflow_Detail_UsePhase_B).copy()
            ProductOutflow[:,Sector_reb_rge,mS,mR]      = np.einsum('tcBr->tB',Outflow_Detail_UsePhase_B).copy()
            
            # Include renovation of reb. not anymore. now just time dependent
            RECC_System.ParameterDict['3_MC_RECC_Buildings_t'].Values[:,:,:,:,:,mS] = np.einsum('cmBr,t->mBrct',RECC_System.ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,:,:,:,mS],np.ones(Nt)) # mBrctS
            RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,:,:,:] = np.einsum('cBVnr,trcB->cBVnrt',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:SwitchTime,:,:,:,:,mS],np.ones((Nt,Nr,Nc-Nt+1,NB))) # cBVnrt
            # Add values for future age-cohorts, convert from useful to final energy, expand from 'all' to specific energy carriers
            RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[SwitchTime-1::,:,:,:,:,:]   = np.einsum('Vrnt,Vrnt,cBVr,t->cBVnrt',ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[:,mR,:,:,:,mS],ParameterDict['3_SHA_EnergyCarrierSplit_Buildings_uf'].Values[:,mR,:,:,:,mS],RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchTime-1::,:,:,all_loc,:,mS],np.ones(Nt))
            # Split energy into different carriers for historic age-cohorts:
            RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,:,:,:]     = np.einsum('Vrnt,cBVrt->cBVnrt',RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values[:,mR,:,:,:], RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,all_loc,:,:]) 
             
            # Prepare parameters:        
            Par_RECC_MC_Nr = np.zeros((Nc,Nm,Ng,Nr,NS,Nt))  # Unit:  kg/m².
            Par_RECC_MC_Nr[:,:,Sector_reb_rge,:,mS,:] = np.einsum('mBrct->Bcmrt',RECC_System.ParameterDict['3_MC_RECC_Buildings_t'].Values[:,:,:,:,:,mS])
            # historic element composition of materials:
            Par_Element_Composition_of_Materials_m   = np.zeros((Nc,Nm,Ne)) # Unit: 1. Aspects: cme, produced in age-cohort c. Applies to new manufactured goods.
            Par_Element_Composition_of_Materials_m[0:Nc-Nt+1,:,:] = np.einsum('c,me->cme',np.ones(Nc-Nt+1),RECC_System.ParameterDict['3_MC_Elements_Materials_ExistingStock'].Values)
            # For future age-cohorts, the total is known but the element breakdown of this parameter will be updated year by year in the loop below.
            Par_Element_Composition_of_Materials_m[:,:,0] = 1 # element 0 is 'all', for which the mass share is always 100%.
            # future element composition of materials inflow use phase (mix new and reused products)
            Par_Element_Composition_of_Materials_c   = np.zeros((Nt,Nm,Ne)) # cme, produced in age-cohort c. Applies to new manufactured goods.
            # Element composition of material in the use phase
            Par_Element_Composition_of_Materials_u   = Par_Element_Composition_of_Materials_m.copy() # cme
            # Determine total element composition of products (c: age-cohort), needs to be updated for future age-cohorts, is done below after material cycle computation.
            Par_3_MC_Stock_ByElement_Nr = np.einsum('cmgrt,cme->tcrgme',Par_RECC_MC_Nr[:,:,:,:,mS,:],Par_Element_Composition_of_Materials_m) # Unit: vehicles: kg/item, buildings: kg/m².
            
            Mylog.info('Translate total flows into individual materials and elements, for 2015 and historic age-cohorts.')
            
            # convert product stocks and flows to material stocks and flows, only for chemical element position 'all':
            # Stock elemental composition, historic for each element and for future years: 'all' elements only
            RECC_System.StockDict['S_2'].Values[:,:,:,Sector_reb_rge,:,:] = \
                np.einsum('tcrBme,tcBr->tcrBme',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_reb_rge,:,:],Stock_Detail_UsePhase_B)/1000   # Indices='t,c,r,B,m'
            # Outflow, 'all' elements only:
            RECC_System.FlowDict['F_2_0'].Values[:,:,:,Sector_reb_rge,:,0] = \
                np.einsum('Btcrm,tcBr->Btcrm',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_reb_rge,:,0],Outflow_Detail_UsePhase_B)/1000 # all elements, Indices='t,c,r,B,m'
            # Inflow of renovation material as stock multiplied with change in material composition:
            # inflow of materials in new products
            for mmt in range(0,Nt):
                F_1_2_new[mmt,:,Sector_reb_rge,:,0] = np.einsum('Br,Brm->Brm',Inflow_Detail_UsePhase_B[mmt,:,:],Par_3_MC_Stock_ByElement_Nr[mmt,SwitchTime+mmt-1,:,Sector_reb_rge,:,0])/1000
            RECC_System.FlowDict['F_1_2'].Values[:,:,Sector_reb_rge,:,0]   = np.einsum('Btrm->Btrm',F_1_2_new[:,:,Sector_reb_rge,:,0])  
                
            Mylog.info('Calculate material stocks and flows, material cycles, determine elemental composition.')
            
            # Prepare parameter for manufactoring loss
            Par_FabYieldLoss = RECC_System.ParameterDict['4_PY_Manufacturing'].Values.copy() 
            Par_FabYieldLoss        = Par_FabYieldLoss #mgto
            Par_FabYieldLoss_total  = np.einsum('mgto->mgto',Par_FabYieldLoss)
            Divisor                 = 1-Par_FabYieldLoss_total
            Par_FabYield_total_inv  = np.divide(1, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # mgto
            
            for t in tqdm(range(1,Nt), unit=' years'):  # 1: 2016
                CohortOffset = t + Nc-Nt # index of current age-cohort. 
                # Split outflow in material and elements
                RECC_System.FlowDict['F_2_0'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,:] = \
                    np.einsum('Bcrme,cBr->Bcrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge,:,:],Outflow_Detail_UsePhase_B[t,0:CohortOffset,:,:])/1000 # All elements.
                # Aggregated consumption. Production is calculated as a market
                F_1_2_aggregated[t,0,:,:,0]                   = np.einsum('rgm->gm',RECC_System.FlowDict['F_1_2'].Values[t,:,:,:,0])
                # Calculate total input for manufacturing, based on manufacturing yields
                Manufacturing_Input_m_adj    = np.einsum('mg,gm->m', Par_FabYield_total_inv[:,:,t,0],F_1_2_aggregated[t,0,:,:,0]).copy()
                Manufacturing_Input_gm_adj   = np.einsum('mg,gm->gm',Par_FabYield_total_inv[:,:,t,0],F_1_2_aggregated[t,0,:,:,0]).copy()
                # split manufacturing material input into different products g:
                Manufacturing_Input_Split_gm = np.einsum('gm,m->gm', Manufacturing_Input_gm_adj, np.divide(1, Manufacturing_Input_m_adj, out=np.zeros_like(Manufacturing_Input_m_adj), where=Manufacturing_Input_m_adj!=0))
                # Total manufactyring inputs corrspond to production requirements
                PrimaryProductionDemand = Manufacturing_Input_m_adj
                RECC_System.FlowDict['F_0_1'].Values[t,:,:]        = np.einsum('m,me->me',PrimaryProductionDemand,RECC_System.ParameterDict['3_MC_Elements_Materials_Primary'].Values)
                
                # Element composition of material flows:         
                Manufacturing_Input_me_final                       = RECC_System.FlowDict['F_0_1'].Values[t,:,:]
                Manufacturing_Input_gme_final                      = np.einsum('gm,me->gme',Manufacturing_Input_Split_gm,Manufacturing_Input_me_final)
                Element_Material_Composition_Manufacturing         = msf.DetermineElementComposition_All_Oth(Manufacturing_Input_me_final)  
                Par_Element_Composition_of_Materials_m[t+115,:,:]  = Element_Material_Composition_Manufacturing.copy()
                Par_Element_Composition_of_Materials_u[t+115,:,:]  = Element_Material_Composition_Manufacturing.copy()
    
                # Calculate manufacturing output, at global level only, by elements
                F_1_2_aggregated[t,0,:,:,:]    = np.einsum('me,gm->gme', Element_Material_Composition_Manufacturing ,F_1_2_aggregated[t,0,:,:,0])
                # Calculate material composition of product consumption
                Throughput_FinalGoods_me                           = F_1_2_aggregated[t,0,:,:,:].sum(axis =0) 
                Element_Material_Composition_cons                  = msf.DetermineElementComposition_All_Oth(Throughput_FinalGoods_me)
                Par_Element_Composition_of_Materials_c[t,:,:]      = Element_Material_Composition_cons.copy()
                
                # Calculate manufacturing scrap 
                RECC_System.FlowDict['F_1_0'].Values[t,:,:]     = np.einsum('gme,mg->me',Manufacturing_Input_gme_final,Par_FabYieldLoss[:,:,t,0]) 
                
                # update mat. composition by element for current year and latest age-cohort
                Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_reb_rge,:,:]   = np.einsum('me,Bmr->Brme',Par_Element_Composition_of_Materials_c[t,:,:],Par_RECC_MC_Nr[SwitchTime+t-1,:,Sector_reb_rge,:,mS,t])
                # Determine element breakdown of inflow and renovation material
                RECC_System.FlowDict['F_1_2'].Values[t,:,Sector_reb_rge,:,:]   = \
                    np.einsum('me,Brm->Brme',Par_Element_Composition_of_Materials_c[t,:,:],RECC_System.FlowDict['F_1_2'].Values[t,:,Sector_reb_rge,:,0]) # all elements, Indices='t,r,B,m,e'
                # Value of Par_3_MC_Stock_ByElement_Nr for current year and all previous age-cohorts c < t need to be updated (As they might change due to the adding of recycling materials in the current year. Currently not included)
                # Determine the element material composition at the end of last year, as weighting factor for existing stock
                Divisor  = np.einsum('Bcrm,e->Bcrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge,:,0],np.ones(Ne))
                Par_ElementComposition_LastYear = np.divide(Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge,:,:],Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) #Bcrme
                # Compile all materials present in stock broken down by element:
                # Here, The materials present in stock consist of the current products in stock * their element composition of last year plus the inflow of renovation material with this years material production element composition.
                StockMat = np.einsum('Bcrm,Bcrme->Bcrme',RECC_System.StockDict['S_2'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,0] ,Par_ElementComposition_LastYear)
                Divisor  = np.einsum('Bcrm,e->Bcrme',StockMat[:,:,:,:,0],np.ones(Ne))
                # Calculate product element composition of latest age-cohort from total materials by element:
                Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset,:,Sector_reb_rge,:,:]  = np.einsum('Bcmr,Bcrme->Bcrme',Par_RECC_MC_Nr[0:CohortOffset,:,Sector_reb_rge,:,mS,t],np.divide(StockMat,Divisor, out=np.zeros_like(Divisor), where=Divisor!=0))
                # Update stock: break down material into elements:                
                RECC_System.StockDict['S_2'].Values[t,0:CohortOffset +1,:,Sector_reb_rge,:,:] = \
                    np.einsum('Bcrme,cBr->Bcrme',Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset +1,:,Sector_reb_rge,:,:],Stock_Detail_UsePhase_B[t,0:CohortOffset +1,:,:])/1000
                    
                # Calculate stock changes
                RECC_System.StockDict['dS_2'].Values[t,:,:,:,:,:]     = RECC_System.StockDict['S_2'].Values[t,:,:,:,:,:]    - RECC_System.StockDict['S_2'].Values[t-1,:,:,:,:,:]
                RECC_System.StockDict['dS_0'].Values[t,:] = \
                    np.einsum('crgme->e',RECC_System.FlowDict['F_2_0'].Values[t,:,:,:,:,:]) + \
                    np.einsum('me->e',RECC_System.FlowDict['F_1_0'].Values[t,:,:]) -\
                    np.einsum('me->e',RECC_System.FlowDict['F_0_1'].Values[t,:,:])
                
            # Flows calculations concluded. Check whether flow value arrays match their indices, etc.
            RECC_System.Consistency_Check() 
            # Determine Mass Balance
            Bal = RECC_System.MassBalance()
            BalAbs = np.abs(Bal).sum()
            Mylog.info('Total mass balance deviation (np.abs(Bal).sum() for socioeconomic scenario ' + SName + ' and RE scenario ' + RName + ': ' + str(BalAbs) + ' Mt.')                    
            
            
            ##########################################################
            #    Section 6) Post-process RECC model solution         #
            ##########################################################  
            Mylog.info('## Compile results array.')          
            # All GHG and material flows/indicators in Mt/yr ot t/cap/yr if per capita. All energy flows in TJ/yr.
            # All energy flows are _final_ energy unless otherwise indicated by variable name.
            
            # Assign material flows to system variables
            SysVar_Material_Demand         = np.einsum('trgm->tm', RECC_System.FlowDict['F_1_2'].Values[:,:,:,:,0])
            SysVar_Material_Demand_pc      = SysVar_Material_Demand / np.einsum('t,m->tm', Population[:,mS,mR], np.ones((Nm)) )
            SysVar_Material_Outflow        = np.einsum('tcrgm->tm', RECC_System.FlowDict['F_2_0'].Values[:,:,:,:,:,0] )
            SysVar_Material_Stock          = np.einsum('tcrgm->tm', RECC_System.StockDict['S_2'].Values[:,:,:,:,:,0] )
            SysVar_Material_Stock_pc       = SysVar_Material_Stock / np.einsum('t,m->tm', Population[:,mS,mR], np.ones((Nm)) )
            SysVar_Material_Production     = RECC_System.FlowDict['F_0_1'].Values[:,:,0].copy()
            SysVar_Material_Production_pc  = SysVar_Material_Production / np.einsum('t,m->tm', Population[:,mS,mR], np.ones((Nm)) )
            SysVar_Manufacturing_Scrap     = RECC_System.FlowDict['F_1_0'].Values[:,:,0].copy()
            
            SysVar_Product_Inflow          = np.einsum('tBr->tB',  Inflow_Detail_UsePhase_B)
            SysVar_Product_Outflow         = np.einsum('tcBr->tB', Outflow_Detail_UsePhase_B)
            SysVar_Product_Stock_total     = np.einsum('tcBr->t',   Stock_Detail_UsePhase_B)
            SysVar_Product_Stock_by_type   = np.einsum('tcBr->tB',  Stock_Detail_UsePhase_B)
    
            # Calculate use phase energy consumption
            SysVar_EnergyDemand_UsePhase_ByService          = np.einsum('cBVnrt,tcBVr,tcBr->tV',  ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values,RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],Stock_Detail_UsePhase_B)
            SysVar_EnergyDemand_UsePhase_ByService_r        = np.einsum('cBVnrt,tcBVr,tcBr->tVr', ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values,RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],Stock_Detail_UsePhase_B)
            SysVar_EnergyDemand_UsePhase_ByEnergyCarrier    = np.einsum('cBVnrt,tcBVr,tcBr->tn',  ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values,RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],Stock_Detail_UsePhase_B)
            SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_r  = np.einsum('cBVnrt,tcBVr,tcBr->tnr', ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values,RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],Stock_Detail_UsePhase_B)
            SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_EL = SysVar_EnergyDemand_UsePhase_ByEnergyCarrier[:,Electric_loc]
            # Calculate manufacturing energy demand
            SysVar_EnergyDemand_Manufacturing_ByEnergyCarrier   = np.einsum('Bn,tBr->tn', RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_reb_rge,:,110,-1],Inflow_Detail_UsePhase_B)  # conversion factor: 1, as MJ/m² = TJ/Million m²
            SysVar_EnergyDemand_Manufacturing_ByEnergyCarrier_r = np.einsum('Bn,tBr->tnr',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_reb_rge,:,110,-1],Inflow_Detail_UsePhase_B)  # conversion factor: 1, as MJ/m² = TJ/Million m²
            # Total energy demand
            SysVar_EnergyDemand_Total_ByEnergyCarrier = SysVar_EnergyDemand_UsePhase_ByEnergyCarrier + SysVar_EnergyDemand_Manufacturing_ByEnergyCarrier
            
            # Unit: Mt/yr. 1 kg/MJ = 1kt/TJ
            # Calculate impacts for energy supply during use phase: direct (combustion of carriers), indirect and total.
            SysExt_DirectImpacts_EnergySupply_UsePhase_ByEnergyCarrier     = 0.001 * np.einsum('Xn,xX,tn->xtn',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values, RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values, SysVar_EnergyDemand_UsePhase_ByEnergyCarrier)
            SysExt_IndirectImpacts_EnergySupply_UsePhase_ByEnergyCarrier   = 0.001 * np.einsum('nxrt,tnr->xtn',   RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_r) 
            SysExt_TotalImpacts_EnergySupply_UsePhase_ByEnergyCarrier      = SysExt_DirectImpacts_EnergySupply_UsePhase_ByEnergyCarrier + SysExt_IndirectImpacts_EnergySupply_UsePhase_ByEnergyCarrier
            # Split indirect impacts between electricity and others
            SysExt_IndirectImpacts_EnergySupply_UsePhase_EL   = 0.001 * np.einsum('xrt,tr->xt',  RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[Electric_loc,:,:,:,mR], SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_r[:,Electric_loc,:])
            SysExt_IndirectImpacts_EnergySupply_UsePhase_OtherThanEL = np.einsum('xtn->xt',SysExt_IndirectImpacts_EnergySupply_UsePhase_ByEnergyCarrier) - SysExt_IndirectImpacts_EnergySupply_UsePhase_EL
            # Similar to above but, split by service
            SysExt_DirectImpacts_EnergySupply_UsePhase_ByService     = 0.001 * np.einsum('Xn,xX,tV->xtV',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values, RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values, SysVar_EnergyDemand_UsePhase_ByService)
            SysExt_IndirectImpacts_EnergySupply_UsePhase_ByService   = 0.001 * np.einsum('nxrt,tVr->xtV',   RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByService_r) 
            SysExt_TotalImpacts_EnergySupply_UsePhase_ByService      = SysExt_DirectImpacts_EnergySupply_UsePhase_ByService + SysExt_IndirectImpacts_EnergySupply_UsePhase_ByService
            # Similar to above, but aggregated and not split
            SysExt_DirectImpacts_EnergySupply_UsePhase     = 0.001 * np.einsum('Xn,xX,tn->xt', RECC_System.ParameterDict['6_PR_DirectEmissions'].Values, RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values, SysVar_EnergyDemand_UsePhase_ByEnergyCarrier)
            SysExt_IndirectImpacts_EnergySupply_UsePhase   = 0.001 * np.einsum('nxrt,tnr->xt', RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,:,:,mR],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_r) 
            SysExt_TotalImpacts_EnergySupply_UsePhase      = SysExt_DirectImpacts_EnergySupply_UsePhase + SysExt_IndirectImpacts_EnergySupply_UsePhase
            
            # Calculate impacts for manufacturing
            SysExt_IndirectImpacts_EnergySupply_Manufacturing  = 0.001 * np.einsum('nxt,tnr->xt',   RECC_System.ParameterDict['4_PE_ProcessExtensions_EnergyCarriers_MJ_r'].Values[:,:,0,:,mR], SysVar_EnergyDemand_Manufacturing_ByEnergyCarrier_r)
            SysExt_DirectImpacts_Manufacturing                 = 0.001 * np.einsum('Xn,xX,tnr->xt', RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values, SysVar_EnergyDemand_Manufacturing_ByEnergyCarrier_r)
            SysExt_TotalImpacts_Manufacturing                  = SysExt_DirectImpacts_Manufacturing + SysExt_IndirectImpacts_EnergySupply_Manufacturing
            
            # Calculate impacts for material production
            if Nr>1:
                #Par_4_PE_ProcessExtensions_Materials_average = np.mean(RECC_System.ParameterDict['4_PE_ProcessExtensions_Materials_extended'].Values, axis=2) # mxrt->mxt
                #SysExt_TotalImpacts_Production            = np.einsum('mx,tm->xt',  Par_4_PE_ProcessExtensions_Materials_average[:,:,0],  RECC_System.FlowDict['F_0_1'].Values[:,:,0])
                #SysExt_TotalImpacts_Production_ByMaterial = np.einsum('mx,tm->xtm', Par_4_PE_ProcessExtensions_Materials_average[:,:,0],RECC_System.FlowDict['F_0_1'].Values[:,:,0])
                Par_4_PE_ProcessExtensions_Materials_average = np.mean(RECC_System.ParameterDict['4_PE_ProcessExtensions_Materials_extended'].Values, axis=2) # mxrRt->mxRt
                SysExt_TotalImpacts_Production            = np.einsum('mxt,tm->xt',  Par_4_PE_ProcessExtensions_Materials_average[:,:,mR,:],  RECC_System.FlowDict['F_0_1'].Values[:,:,0])
                SysExt_TotalImpacts_Production_ByMaterial = np.einsum('mxt,tm->xtm', Par_4_PE_ProcessExtensions_Materials_average[:,:,mR,:],RECC_System.FlowDict['F_0_1'].Values[:,:,0])
            else:
                SysExt_TotalImpacts_Production            = np.einsum('mxt,tm->xt',  RECC_System.ParameterDict['4_PE_ProcessExtensions_Materials_extended'].Values[:,:,0,mR,:],RECC_System.FlowDict['F_0_1'].Values[:,:,0])
                SysExt_TotalImpacts_Production_ByMaterial = np.einsum('mxt,tm->xtm', RECC_System.ParameterDict['4_PE_ProcessExtensions_Materials_extended'].Values[:,:,0,mR,:],RECC_System.FlowDict['F_0_1'].Values[:,:,0]) 
                
            # Calculate total impacts
            SysExt_TotalImpacts = SysExt_TotalImpacts_EnergySupply_UsePhase + SysExt_TotalImpacts_Manufacturing + SysExt_TotalImpacts_Production
            
            # Calculate wood C uptake by wood, as a carbon sink
            SysExt_C_Uptake = -1 * np.einsum('tm->t', RECC_System.FlowDict['F_0_1'].Values[:,:,Carbon_loc]) # negative sign because emissions are measured in X_0 direction.
            # SysExt_CO2_Uptake = -1 * 44/12 * np.einsum('tm->t', RECC_System.FlowDict['F_0_1'].Values[:,:,Carbon_loc]) # negative sign because emissions are measured in X_0 direction.
            # Total GHG
            SysExt_TotalGHG_wo_uptake = SysExt_TotalImpacts[GWP100_loc,:]
            # SysExt_TotalGHG_w_uptake  = SysExt_TotalGHG_wo_uptake + SysExt_CO2Uptake
            
            # Compile results arrays
            MaterialDemand[:,:,mS,mR]                              = SysVar_Material_Demand
            MaterialDemand_pc[:,:,mS,mR]                           = SysVar_Material_Demand_pc
            DemolitionWaste[:,:,mS,mR]                             = SysVar_Material_Outflow
            MaterialStock[:,:,mS,mR]                               = SysVar_Material_Stock
            MaterialStock_pc[:,:,mS,mR]                            = SysVar_Material_Stock_pc
            MaterialProduction[:,:,mS,mR]                          = SysVar_Material_Production
            MaterialProduction_pc[:,:,mS,mR]                       = SysVar_Material_Production_pc
            ManufacturingScrap[:,:,mS,mR]                          = SysVar_Manufacturing_Scrap
            EnergyCons_UsePhase_byEnergyCarrier[:,:,mS,mR]         = SysVar_EnergyDemand_UsePhase_ByEnergyCarrier
            EnergyCons_UsePhase_byEnergyCarrier_EL[:,mS,mR]        = SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_EL
            EnergyCons_UsePhase_byService[:,:,mS,mR]               = SysVar_EnergyDemand_UsePhase_ByService
            EnergyCons_Manufacturing_byEnergyCarrier[:,:,mS,mR]    = SysVar_EnergyDemand_Manufacturing_ByEnergyCarrier
            EnergyCons_Total_byEnergyCarrier[:,:,mS,mR]            = SysVar_EnergyDemand_Total_ByEnergyCarrier
            Impacts_UsePhase_byEnergyCarrier[:,:,:,mS,mR]          = SysExt_TotalImpacts_EnergySupply_UsePhase_ByEnergyCarrier
            Impacts_UsePhase_byEnergyCarrier_direct[:,:,:,mS,mR]   = SysExt_DirectImpacts_EnergySupply_UsePhase_ByEnergyCarrier
            Impacts_UsePhase_byEnergyCarrier_indirect[:,:,:,mS,mR] = SysExt_IndirectImpacts_EnergySupply_UsePhase_ByEnergyCarrier
            Impacts_UsePhase_indir_EL[:,:,mS,mR]                   = SysExt_IndirectImpacts_EnergySupply_UsePhase_EL
            Impacts_UsePhase_indir_otherThanEL[:,:,mS,mR]          = SysExt_IndirectImpacts_EnergySupply_UsePhase_OtherThanEL
            Impacts_UsePhase_byService[:,:,:,mS,mR]                = SysExt_TotalImpacts_EnergySupply_UsePhase_ByService
            Impacts_UsePhase[:,:,mS,mR]                            = SysExt_TotalImpacts_EnergySupply_UsePhase
            Impacts_UsePhase_direct[:,:,mS,mR]                     = SysExt_DirectImpacts_EnergySupply_UsePhase
            Impacts_UsePhase_indirect[:,:,mS,mR]                   = SysExt_IndirectImpacts_EnergySupply_UsePhase
            Impacts_Manufacturing[:,:,mS,mR]                       = SysExt_TotalImpacts_Manufacturing
            Impacts_Production[:,:,mS,mR]                          = SysExt_TotalImpacts_Production
            Impacts_Production_byMaterials[:,:,:,mS,mR]            = SysExt_TotalImpacts_Production_ByMaterial
            Impacts_Total[:,:,mS,mR]                               = SysExt_TotalImpacts
            C_Uptake[:,mS,mR]                                      = SysExt_C_Uptake
            # CO2_Uptake[:,mS,mR]                                    = SysExt_CO2_Uptake
            GHG_total_wo_uptake[:,mS,mR]                           = SysExt_TotalGHG_wo_uptake
            # GHG_total_w_uptake[:,mS,mR]                            = SysExt_TotalGHG_w_uptake
            
            # Extract calibration for SSP1:
            if mS == 1:
                E_Calib_Buildings = np.einsum('tnr->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_r[:,:-1,:])
                
            # Determine exit flags            
            ExitFlags['Positive_Inflow_F1_2_R32_SSP_'  + str(mS) + '_RCP_' + str(mR)] = np.isclose(RECC_System.FlowDict['F_1_2'].Values.min(),0, IsClose_Remainder_Small)
            ExitFlags['Positive_Outflow_F2_0_R32_SSP_' + str(mS) + '_RCP_' + str(mR)] = np.isclose(RECC_System.FlowDict['F_2_0'].Values.min(),0, IsClose_Remainder_Small)  
                    
            Mylog.debug('----\n')
            
                    
    #######################################
    #   Section 7) Export results         #
    #######################################
    Mylog.info('## Export results.')
    
    # 7.1) Exit flags
    Mylog.info('Check data and results for boundary constraints and plausibility. Exit flags.')
    Mylog.info('Model input')          
    ExitFlags['3_SHA_TypeSplit_Buildings_min'] = ParameterDict['3_SHA_TypeSplit_Buildings'].Values.min() >= 0
    ExitFlags['3_SHA_TypeSplit_Buildings_max'] = ParameterDict['3_SHA_TypeSplit_Buildings'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_Buildings_sum'] = np.isclose(ParameterDict['3_SHA_TypeSplit_Buildings'].Values.sum(),Nr*Nt*NS, IsClose_Remainder_Large)
    
    Mylog.info('Model exit flags:')
    for key in ExitFlags:
        Mylog.info(key + ': ' + str(ExitFlags[key]))
    
    # 7.2) Write excel results workbooks
    book = openpyxl.Workbook() # Model results in iamc style (row: specifier, columns: years)
    ws1 = book.active
    ws1.title = 'Cover'
    ws1.cell(row=3, column=2).value = 'ScriptConfig'
    ws1.cell(row=3, column=2).font = openpyxl.styles.Font(bold=True)
    m = 4
    for x in sorted(ScriptConfig.keys()):
        ws1.cell(row=m, column=2).value = x
        ws1.cell(row=m, column=3).value = ScriptConfig[x]
        m +=1
    
    ws2 = book.create_sheet('Model_Results')
    ColLabels = ['Indicator','Unit','Region','System_location','RE scen','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws2.cell(row=1, column=m+1).value = ColLabels[m]
        ws2.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws2.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws2.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    
    # population
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Population[:,:,:],2,len(ColLabels),'Population','million',ScriptConfig['RegionalScope'],'P (population)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # ProductStock_byType
    for g in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,ProductStock_byType[:,g,:,:],newrowoffset,len(ColLabels),'Product stock: ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[g],'million m2',ScriptConfig['RegionalScope'],'S_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # EnergyCons_UsePhase_byService
    for V in range(0,NV):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UsePhase_byService[:,V,:,:],newrowoffset,len(ColLabels),'Energy consumption, use phase: ' + IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items[V],'TJ/yr',ScriptConfig['RegionalScope'],'E_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # EnergyCons_UsePhase_byEnergyCarrier
    for n in range(0,Nn-1):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UsePhase_byEnergyCarrier[:,n,:,:],newrowoffset,len(ColLabels),'Energy consumption, use phase: ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[n],'TJ/yr',ScriptConfig['RegionalScope'],'E_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Impacts_UsePhase_direct
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_direct[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, use phase, direct','Mt CO2-eq/yr',ScriptConfig['RegionalScope'],'S_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Impacts_UsePhase_indirect
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_indirect[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, use phase, indirect','Mt CO2-eq/yr',ScriptConfig['RegionalScope'],'S_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Total impacts, production, by material
    for m in range(Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Production_byMaterials[GWP100_loc,:,m,:,:],newrowoffset,len(ColLabels),'GHG emissions, material production: ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[m] ,'Mt CO2-eq/yr',ScriptConfig['RegionalScope'],'F_0_1','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Total impacts, use phase, by energy carrier (direct + indirect)
    for n in range(Nn-1):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_byEnergyCarrier[GWP100_loc,:,n,:,:],newrowoffset,len(ColLabels),'GHG emissions, use phase (dir+indir): ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[n],'Mt CO2-eq/yr',ScriptConfig['RegionalScope'],'S_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Total impacts, use phase, by service
    for V in range(NV):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_UsePhase_byService[GWP100_loc,:,V,:,:],newrowoffset,len(ColLabels),'GHG emissions, use phase: ' + IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items[V],'Mt CO2-eq/yr',ScriptConfig['RegionalScope'],'S_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Total impacts (use phase + manu + prod)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Total[GWP100_loc,:,:,:],newrowoffset,len(ColLabels),'GHG emissions, system-wide','Mt CO2-eq/yr',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for x in range(Nx):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Impacts_Total[x,:,:,:],newrowoffset,len(ColLabels),'Impact, system-wide: ' + IndexTable.Classification[IndexTable.index.get_loc('Environmental pressure')].Items[x]  ,'GWP: Mt CO2-eq/yr, LOP:1000 km2, WCP: billions m3',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # CO2_Uptake
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,C_Uptake[:,:,:],newrowoffset,len(ColLabels),'Carbon uptake','Mt/yr',ScriptConfig['RegionalScope'],'F_0_1','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # final material consumption
    for m in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,MaterialDemand[:,m,:,:],newrowoffset,len(ColLabels),'Final consumption of materials: ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[m],'Mt/yr',ScriptConfig['RegionalScope'],'F_1_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # material production
    for m in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,MaterialProduction[:,m,:,:],newrowoffset,len(ColLabels),'Production of materials: ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[m],'Mt/yr',ScriptConfig['RegionalScope'],'F_0_1','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # demolition waste
    for m in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,DemolitionWaste[:,m,:,:],newrowoffset,len(ColLabels),'Demolition waste: ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[m],'Mt/yr',ScriptConfig['RegionalScope'],'F_2_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # stock of materials 
    for m in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,MaterialStock[:,m,:,:],newrowoffset,len(ColLabels),'Material stock: ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[m],'Mt',ScriptConfig['RegionalScope'],'S_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for m in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,MaterialStock_pc[:,m,:,:],newrowoffset,len(ColLabels),'Material stock, per capita: ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[m],'t',ScriptConfig['RegionalScope'],'S_2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    # # Export total material stock
    # pd_xlsx_writer = pd.ExcelWriter(os.path.join(ProjectSpecs_Path_Result,'ODYM_RECC_Additional_Results_'+ ScriptConfig['Current_UUID'] + '.xlsx'), engine="xlsxwriter")
    # ColIndex       = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items
    # RowIndex        = pd.MultiIndex.from_product([IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items,IndexTable.Classification[IndexTable.index.get_loc('ResidentialBuildings')].Items], names=('Region','Stock_Item'))
    # DF_matstock2015 = pd.DataFrame(np.einsum('mBr->rBm',TotalMaterialStock_2015_reb).reshape(Nr*NB,Nm), index=RowIndex, columns=ColIndex)
    # DF_matstock2015.to_excel(pd_xlsx_writer, sheet_name="RECC_matstock_2015_reb_Mt", merge_cells=False)
    # pd_xlsx_writer.close()
       
    # book2 = openpyxl.load_workbook(os.path.join(ProjectSpecs_Path_Result,'ODYM_RECC_Additional_Results_'+ ScriptConfig['Current_UUID'] + '.xlsx')) # Export other model results, calibration values, flags, etc.
    # cover_sheet = book2.create_sheet('Cover')
    # cover_sheet.cell(row=3, column=2).value = 'ScriptConfig'
    # cover_sheet.cell(row=3, column=2).font = openpyxl.styles.Font(bold=True)
    # m = 4
    # for x in sorted(ScriptConfig.keys()):
    #     cover_sheet.cell(row=m, column=2).value = x
    #     cover_sheet.cell(row=m, column=3).value = ScriptConfig[x]
    #     m +=1       
        
    # reb_Sheet = book2.create_sheet('residential buildings')
    # reb_Sheet.cell(1,2).value = '2015 post calibration values, by model region'
    # reb_Sheet.cell(1,2).font  = openpyxl.styles.Font(bold=True)
    # reb_Sheet.cell(2,2).value = 'region'
    # reb_Sheet.cell(2,2).font  = openpyxl.styles.Font(bold=True)
    # m=2
    # for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
    #     reb_Sheet.cell(m+1,2).value = Rname
    #     reb_Sheet.cell(m+1,2).font  = openpyxl.styles.Font(bold=True)        
    #     m+=1
    
    # # pC stock values
    # reb_Sheet.cell(2,3).value = '2015 use phase energy consumption, across all building types and energy standards, by model region. Unit: TJ/yr. Value for SSP1.'
    # reb_Sheet.cell(2,3).font  = openpyxl.styles.Font(bold=True)      
    # m=2
    # for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region_Focus')].Items:
    #     reb_Sheet.cell(m+1,3).value = E_Calib_Buildings[0,m-2]
    #     m+=1
    
    # 7.3) Export to Excel
    Mylog.info('### Export to Excel')
    # Export list data
    book.save( os.path.join(ProjectSpecs_Path_Result,'ODYM_RECC_ModelResults_' + ScriptConfig['Current_UUID'] + '.xlsx'))
    # book2.save(os.path.join(ProjectSpecs_Path_Result,'ODYM_RECC_Additional_Results_'+ ScriptConfig['Current_UUID'] + '.xlsx'))
    
    # 7.4) Export .dat file
    Mylog.info('### Export to .dat')
    # write dat file for results
    OutputDict['Population'] = Population
    OutputDict['Material demand'] = MaterialDemand
    OutputDict['Material demand, per capita'] = MaterialDemand_pc    
    OutputDict['Demolition waste'] = DemolitionWaste      
    OutputDict['Material stock'] = MaterialStock        
    OutputDict['Material stock, per capita'] = MaterialStock_pc      
    OutputDict['Material production'] = MaterialProduction    
    OutputDict['Material production, per capita'] = MaterialProduction_pc 
    OutputDict['Manufacturing scrap'] = ManufacturingScrap    
    OutputDict['Inflow of products'] = ProductInflow       
    OutputDict['Outflow of products'] = ProductOutflow      
    OutputDict['Stock of products'] = ProductStock        
    OutputDict['Stock pf products, by type'] = ProductStock_byType 
    OutputDict['Stock of products, per capita'] = ProductStock_pc     
    OutputDict['Energy consumption, use phase, by carrier'] = EnergyCons_UsePhase_byEnergyCarrier      
    OutputDict['Energy consumption, use phase, electricity'] = EnergyCons_UsePhase_byEnergyCarrier_EL   
    OutputDict['Energy consumption, use phase, by service'] = EnergyCons_UsePhase_byService            
    OutputDict['Energy consumption, manufacturing, by carrier'] = EnergyCons_Manufacturing_byEnergyCarrier 
    OutputDict['Total energy consumption, by carrier'] = EnergyCons_Total_byEnergyCarrier         
    OutputDict['Total impacts, use phase, by carrier'] = Impacts_UsePhase_byEnergyCarrier           
    OutputDict['Total direct impacts, use phase, by carrier'] = Impacts_UsePhase_byEnergyCarrier_direct   
    OutputDict['Total indirect impacts, use phase, by carrier'] = Impacts_UsePhase_byEnergyCarrier_indirect 
    OutputDict['Total direct impacts, use phase, electricity'] = Impacts_UsePhase_indir_EL                 
    OutputDict['Total direct impacts, use phase, other than el.'] = Impacts_UsePhase_indir_otherThanEL        
    OutputDict['Total impacts, use phase, by service'] = Impacts_UsePhase_byService                
    OutputDict['Total impacts, use phase'] = Impacts_UsePhase                          
    OutputDict['Total direct impacts, use phase'] = Impacts_UsePhase_direct                   
    OutputDict['Total indirect impacts, use phase'] = Impacts_UsePhase_indirect                 
    OutputDict['Total impacts, manufacturing'] = Impacts_Manufacturing                     
    OutputDict['Total impacts, production'] = Impacts_Production                        
    OutputDict['Total impacts, production, by material'] = Impacts_Production_byMaterials            
    OutputDict['Total impacts'] = Impacts_Total                             
    OutputDict['C uptake'] = C_Uptake        
    # OutputDict['CO2 uptake'] = CO2_Uptake 
    OutputDict['Total GHG, without uptake'] = GHG_total_wo_uptake 
    # OutputDict['Total GHG, with uptake'] = GHG_total_w_uptake  
    
    if ScriptConfig['Save dat']=='True':
        ParFileName = os.path.join(ProjectSpecs_Path_Result,'ODYM_RECC_ModelResults_' + ScriptConfig['Current_UUID'] + '.dat')
        ParFileObject = open(ParFileName,'wb') 
        pickle.dump(OutputDict,ParFileObject)   
        ParFileObject.close()
    
    Mylog.debug('----\n')
    
    ##############################
    #   Section 8) Plots         #
    ##############################
    Mylog.info('## Create plots')
    
    Figurecounter = 1
    # Plot system emissions, by process, stacked.
    MyColorCycle = pylab.cm.gist_earth(np.arange(0,1,0.155)) # select 12 colors from the 'Set1' color map.      
    
    SSPScens   = ['LED','SSP1','SSP2','BAU']
    RCPScens   = ['No climate policy','RCP2.6 energy mix','BAU']
    Area       = ['use phase','use phase, scope 2 (el)','use phase, other indirect','primary material product.','manufact ']     
    
    for mS in range(0,NS): # SSP
        for mR in range(0,NR): # RCP
            fig  = plt.figure(figsize=(8,5))
            ax1  = plt.axes([0.08,0.08,0.85,0.9])
            ProxyHandlesList = []   # For legend     
            # plot area
            ax1.fill_between(np.arange(2015,2061),np.zeros((Nt)), Impacts_UsePhase_direct[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[1,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[1,:])) # create proxy artist for legend
            ax1.fill_between(np.arange(2015,2061),Impacts_UsePhase_direct[GWP100_loc,:,mS,mR], Impacts_UsePhase_direct[GWP100_loc,:,mS,mR] + Impacts_UsePhase_indir_EL[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[2,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[2,:])) # create proxy artist for legend
            ax1.fill_between(np.arange(2015,2061),Impacts_UsePhase_direct[GWP100_loc,:,mS,mR] + Impacts_UsePhase_indir_EL[GWP100_loc,:,mS,mR], Impacts_UsePhase_direct[GWP100_loc,:,mS,mR] + Impacts_UsePhase_indir_EL[GWP100_loc,:,mS,mR] + Impacts_UsePhase_indir_otherThanEL[GWP100_loc,:,mS,mR], linestyle = '-', facecolor = MyColorCycle[3,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[3,:])) # create proxy artist for legend
            ax1.fill_between(np.arange(2016,2061),Impacts_UsePhase_direct[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_indir_EL[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_indir_otherThanEL[GWP100_loc,1::,mS,mR], Impacts_UsePhase_direct[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_indir_EL[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_indir_otherThanEL[GWP100_loc,1::,mS,mR] + Impacts_Production[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor = MyColorCycle[4,:], linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc=MyColorCycle[4,:])) # create proxy artist for legend    
            ax1.fill_between(np.arange(2016,2061),Impacts_UsePhase_direct[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_indir_EL[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_indir_otherThanEL[GWP100_loc,1::,mS,mR] + Impacts_Production[GWP100_loc,1::,mS,mR], Impacts_UsePhase_direct[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_indir_EL[GWP100_loc,1::,mS,mR] + Impacts_UsePhase_indir_otherThanEL[GWP100_loc,1::,mS,mR] + Impacts_Production[GWP100_loc,1::,mS,mR] + Impacts_Manufacturing[GWP100_loc,1::,mS,mR], linestyle = '-', facecolor ='k', linewidth = 0.5)
            ProxyHandlesList.append(plt.Rectangle((0, 0), 1, 1, fc='k')) # create proxy artist for legend    
            # plt.plot(np.arange(2016,2061), GHG_total_w_uptake[1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            # plta = Line2D(np.arange(2016,2061), GHG_total_w_uptake[1::,mS,mR] , linewidth = linewidth[2], color = 'k')
            # ProxyHandlesList.append(plta) # create proxy artist for legend    
            
            plt.title('GHG emissions, stacked by process group, \n' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.', fontsize = 18)
            plt.ylabel(r'Mt of CO$_2$-eq.', fontsize = 18)
            plt.xlabel('Year', fontsize = 18)
            plt.xticks(fontsize=18)
            plt.yticks(fontsize=18)
            plt.legend(handles = reversed(ProxyHandlesList),labels = reversed(Area), shadow = False, prop={'size':14},ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
            ax1.set_xlim([2024, 2060])
            
            plt.show()
            fig_name = 'GWP_TimeSeries_AllProcesses_Stacked_' + ScriptConfig['RegionalScope'] + ', ' + SSPScens[mS] + ', ' + RCPScens[mR] + '.png'
            # include figure in logfile:
            fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
            # comment out to save disk space in archive:
            if ScriptConfig['Save graphs']=='True':
                fig.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
            Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
            Figurecounter += 1
            
    Mylog.debug('----\n')
    
    ##################################
    #   Section 9) Save and exit     #
    ##################################
    Mylog.info('## Finishing')
    Mylog.debug("Converting " + os.path.join(ProjectSpecs_Path_Result, '..', log_filename))
    # everything from here on will not be included in the converted log file
    msf.convert_log(os.path.join(ProjectSpecs_Path_Result, log_filename), 'md')
    Mylog.info('Script is finished. Terminating logging process and closing all log files.')
    Time_End = time.time()
    Time_Duration = Time_End - Time_Start
    Mylog.info('End of simulation: ' + time.asctime())
    Mylog.info('Duration of simulation: %.1f seconds.' % Time_Duration)
    # remove all handlers from logger
    root = log.getLogger()
    root.handlers = []  # required if you don't want to exit the shell
    log.shutdown()
    
    ### 9.1) Create descriptive folder name and rename result folder
    SectList    = eval(ScriptConfig['SectorSelect'])
    DescrString = '__'
    FirstFlag   = True
    for sect in SectList:
        if FirstFlag is True:
            DescrString += sect
            FirstFlag = False
        else:
            DescrString += '_'
            DescrString += sect
    DescrString += '__'        
    
    REStratList = []
    if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'True':
        REStratList.append('MSU')
    if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'True':
        REStratList.append('ULD')
    if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
        REStratList.append('LTE')
    if ScriptConfig['Include_REStrategy_MoreIntenseUse'] == 'True':
        REStratList.append('MIU')
        
    FirstFlag = True
    if len(REStratList) > 0:
        for REStrat in REStratList:
            if FirstFlag is True:
                DescrString += REStrat
                FirstFlag = False
            else:
                DescrString += '_'
                DescrString += REStrat        
        
    ProjectSpecs_Path_Result_New = os.path.join(RECC_Paths.results_path, Name_Scenario + '__' + TimeString + DescrString)
    try:
        os.rename(ProjectSpecs_Path_Result,ProjectSpecs_Path_Result_New)
    except:
        Mylog.info('Folder file not renamed. Acces is denied')
            
    print('done.')
    
    OutputDict['Name_Scenario'] = Name_Scenario + '__' + TimeString + DescrString # return new scenario folder name to ScenarioControl script
    return OutputDict

# # code for script to be run as standalone function
if __name__ == "__main__":
    main()

# The End.
  