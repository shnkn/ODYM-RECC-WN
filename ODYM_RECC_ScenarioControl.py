# -*- coding: utf-8 -*-
"""
Created on Jan 5th, 2020, as copy of RECC_ScenarioControl_V2_2.py

@author: spauliuk
"""

"""

File RECC_ScenarioControl.py

Script that modifies the RECC config file to run a list of scenarios and executes RECC main script for each scenario config.

"""

# Import required libraries:
import os
import openpyxl

import RECC_Paths # Import path file
import ODYM_RECC_Main

#ScenarioSetting, sheet name of RECC_ModelConfig_List.xlsx to be selected:
ScenarioSetting = 'WN1'

# open scenario sheet
ModelConfigListFile  = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,'RECC_ModelConfig_List.xlsx'))
ModelConfigListSheet = ModelConfigListFile[ScenarioSetting]
SheetName = 'Master_RECC_mini'
#Read control lines and execute main model script
ResultFolders = []
Row = 3
# search for script config list entry
while ModelConfigListSheet.cell(Row+1, 3).value != None:
    RegionalScope = ModelConfigListSheet.cell(Row+1, 3).value
    Config = {}
    for m in range(2,11):
        Config[ModelConfigListSheet.cell(3, m+1).value] = ModelConfigListSheet.cell(Row+1, m+1).value
    Row += 1
    # rewrite RECC model config
    mywb = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,'RECC_Config.xlsx'))
    
    sheet = mywb.get_sheet_by_name('Cover')
    sheet['D4'] = SheetName
    sheet = mywb.get_sheet_by_name(SheetName)
    sheet['D7']   = RegionalScope
    sheet['G21']  = Config['RegionSelect']
    # The indices below need to be updated when new parameters are added to the parameter list
    sheet['D85'] = Config['Include_REStrategy_MaterialSubstitution']
    sheet['D86'] = Config['Include_REStrategy_UsingLessMaterialByDesign']
    sheet['D87'] = Config['Include_REStrategy_LifeTimeExtension']
    sheet['D88'] = Config['Include_REStrategy_MoreIntenseUse']
    sheet['D96'] = Config['Save graphs']
    sheet['D97'] = Config['Save dat']
    sheet['D98'] = Config['PlotResolution']
    
    mywb.save(os.path.join(RECC_Paths.data_path,'RECC_Config.xlsx'))

    # run the ODYM-RECC model
    OutputDict = ODYM_RECC_Main.main()
    ResultFolders.append(OutputDict['Name_Scenario'])

# Export ResultFolders:
book = openpyxl.Workbook()
ws1 = book.active
ws1.title = 'ResultFolders'
Fr = 3
for Fname in ResultFolders:
    ws1.cell(row=Fr+1, column=4).value = Fname 
    Fr +=1
book.save(os.path.join(RECC_Paths.results_path,'ResultFolders.xlsx'))   
#
#
